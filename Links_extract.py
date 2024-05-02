from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep 
import openpyxl
from openpyxl import Workbook
from openpyxl import    load_workbook
import requests
from bs4 import BeautifulSoup
#web = webdriver.Chrome()
#Main Category

def Read_main_category():
    wb = Workbook()
    sheet = wb.active
    web.get('https://www.europages.co.uk/companies')
    # click on Category Button
    web.find_element(By.CLASS_NAME,'v-btn__content').click()
    # get all Main Categories in Order
    main_cat=web.find_element(By.CLASS_NAME,'ep-navigation-sectors-list').find_elements(By.TAG_NAME,'a')
    sheet ['A1'] = 'Main Category'
    #Extract All Main Categories To Excel
    for cat in main_cat:
        sheet[f'A{sheet.max_row+1}'] = cat.text.lower()

    wb.save('Main_category.xlsx')
#Sub Categories
def Read_sub_categories():
    wb = load_workbook('Main_category.xlsx')
    sheet = wb.active
    Sub_wb = Workbook()
    Sub_sheet = Sub_wb.active
    Sub_sheet['A1']= 'Main Category'
    Sub_sheet['B1']= 'Sub Category'
    for sub_category in range(2,sheet.max_row):
        #web.get(sheet[f'A{sub_category}'].value)
        web.get(sheet[f'A{sub_category}'].value)
        #read All Sub Categories
        links = web.find_element(By.CLASS_NAME,'ep-business-sectors-two-columns-list').find_elements(By.TAG_NAME,'a')
        sleep(1)
        for link in links:
            max_row = Sub_sheet.max_row+1
            print(link.get_attribute('href'))
            Sub_sheet[f'A{max_row}'] = sheet[f'A{sub_category}'].value
            Sub_sheet[f'B{max_row}'] = link.get_attribute('href')
    
    Sub_wb.save('Sub_Category.xlsx')

# Get The Pagenation For all Sub Category
def Read_pagenations():
    wb = load_workbook('Sub_Category.xlsx')
    r_wb=Workbook()
    r_sh=r_wb.active
    sheet = wb.active
    c=1
    r_sh['A1']= 'Main Category'
    r_sh['B1']= 'Sub Category'
    r_sh['C1']= 'Limit Category'
    for page_limit in range(1,sheet.max_row):
        try : 
            #web.get(sheet[f'B{page_limit+1}'].value)
            c+=1
            #p = web.find_elements(By.TAG_NAME,'ul')[2].text.split('…\n')[1]
            print(p)
            sleep(0.5)
            r_sh[f'A{c}'] = sheet[f'A{page_limit+1}'].value
            r_sh[f'B{c}'] = sheet[f'B{page_limit+1}'].value
            r_sh[f'C{c}'] = p
        except:
            continue

    r_wb.save('test.xlsx')
f = open('error.txt','w')
#get All Companies Links 
company_wb = load_workbook('data.xlsx')
company_sheet=company_wb.active
wb = openpyxl.load_workbook('main.xlsx')
sheet = wb.active
# set Company sheet headers
#get limit 
counter = 2
for l in range(21,sheet.max_row):
    print('Proccess Now >> ', counter)
    counter+=1
    try:
        #get limit
        limit = sheet[f'C{l}'].value
        #get last part of url
        url = sheet[f'B{l}'].value.split('https://www.europages.co.uk/companies/')[1]
        #loop on category with limit
        for pagenation in range(2,int(limit)):
            
            r = requests.get(f'https://www.europages.co.uk/companies/pg-{pagenation}/'+url)
            #print(f'https://www.europages.co.uk/companies/pg-{pagenation}/'+url)
            soup = BeautifulSoup(r.text,'html.parser')
            links=(soup.find_all('a',{'class':'ep-ecard-serp__epage-link'}))
            current_order=0
            for link in links:
                current_order+=1
                companies_row_counter=company_sheet.max_row+1
                data = [sheet[f'A{l}'].value,
                        f'https://www.europages.co.uk/companies/pg-{pagenation}/'+url,
                        sheet[f'C{l}'].value,
                        pagenation,
                        current_order,
                        'https://www.europages.co.uk'+link['href']
                        ]
                company_sheet.append(data)
                #data=[]
    except:
        f.writelines(sheet[f'B{l}'].value)
        f.close()
        continue
    company_wb.save('data.xlsx')


 