from bs4 import BeautifulSoup
import re
import requests
import openpyxl
from openpyxl import Workbook,load_workbook
from time import sleep
import os
wb = load_workbook('1.xlsx')
sheet = wb.active
comp=load_workbook('Scrap.xlsx')
csheet=comp.active
# csheet['A1']='Company Domain'
# csheet['B1']='Company Name'
# csheet['C1']='Company Phones'
# csheet['D1']='Company Website'
# csheet['E1']='Company Facebook'
# csheet['F1']='Company Instagram'
# csheet['G1']='Company Twitter'
# csheet['H1']='Company Linkedin'
# csheet['I1']='Company Describtion'
# csheet['J1']='Company Orginisation'
# csheet['K1']='Company Country'
# csheet['L1']='Company Full Address'
# comp.save('Scrap.xlsx')
facebook = ''
twitter = ''
linkedin = ''
instagram =''
Row_Counter = csheet.max_row+1
Save_counter = 0
#https://www.europages.co.uk/ep-api/v2/epages/HUN055358-00101/phones
for i in range(3,100):
    Save_counter+=1
    r = requests.get(sheet[f'F{i}'].value)
    soup = BeautifulSoup(r.text,'html.parser')
    try:
        site = soup.find('a',{'class':'ep-epages-home-link-card v-card v-sheet v-sheet--outlined theme--light pa-4 ep-epages-home-website-link v-card v-card--link v-sheet theme--light'})['href']
    except:
        site = 'None'
    #print(site)
    try:
        rsite = requests.get(site)
        match = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', rsite.text)
        Email=match.group(0)
    except:
        Email='Null'
    #company_Domain
    company_Domain= soup.select('.v-breadcrumbs__item')[1].text
    #company_Name
    company_Name = soup.select('.v-breadcrumbs__item')[2].text
    #Company describtion
    Describtion= soup.find('div',{'class':"ep-epage-home-description ep-page-epage-home__order-first"}).text
    #Website
    #Orgnaisation
    try:
        org = soup.find('div',{'class':'ep-key-value-list ep-epages-home-business-details__list-column ep-epages-home-business-details__organization'}).text.replace(' ','')
    except:
        org= 'None'
    #Address
    addr = soup.find('dl',{'class':'ep-epages-sidebar__info text--primary pt-4'}).find('dd').find_all('p')
    for add in addr:
        if add.text=='':
            addr = 'None'
        if len(add.text.replace(' ','').replace('\n','').split('-')) >1:
            country = add.text.replace(' ','').replace('\n','').split('-')[1]
    Full_Address = add.text.replace(' ','').replace('\n','')
    company_info = soup.find('h3',{'class':'text-h6 text-sm-h5 pt-10 pb-4'}).text.strip()
    #print(site)
    #Social Media
    try:
        social_media=soup.find('div',{'class':'d-flex align-center mt-4'}).find_all('a')
        for soc in social_media:
            if  'facebook' in soc['href'] :
                facebook = soc['href']
            elif 'twitter' in soc['href']:
                twitter = soc['href']
            elif 'linkedin' in soc['href']:
                linkedin = soc['href']
            elif 'instagram' in soc['href']:
                instagram = soc['href']
    except:
        facebook = 'None'
        twitter = 'None'
        linkedin = 'None'
        instagram = 'None'
    #print('facebook >>',facebook,'\t','instagram >>',instagram,'\t','twitter >>',twitter,'\t','Linkedin >> ',linkedin,'>>'*44)
    try:
        phones = requests.get('https://www.europages.co.uk/ep-api/v2/epages/'+sheet[f'F{i}'].value.replace('https://www.europages.co.uk/','').split('/')[1].replace('.html','')+'/phones').json()['phones'][0]['items']
        phone=''
        if len(phones) ==1:
            phone=phones[0]['number']

        if len(phones) ==2:
            phone=phones[0]['number']+'\n'+phones[1]['number']
        if len(phones) ==3:
            phone=phones[0]['number']+'\n'+phones[1]['number']+'\n'+phones[2]['number']
    except:
        phone='Null'
    #Fill Data
    csheet[f'A{Row_Counter}'] = company_Domain
    csheet[f'B{Row_Counter}'] = company_Name
    csheet[f'C{Row_Counter}'] = phone
    csheet[f'D{Row_Counter}'] = site
    csheet[f'E{Row_Counter}'] = facebook
    csheet[f'F{Row_Counter}'] = instagram
    csheet[f'G{Row_Counter}'] = twitter
    csheet[f'H{Row_Counter}'] = linkedin
    csheet[f'I{Row_Counter}'] = Describtion
    csheet[f'J{Row_Counter}'] = org
    csheet[f'K{Row_Counter}'] = country
    csheet[f'L{Row_Counter}'] = Full_Address
    csheet[f'M{Row_Counter}'] = Email

    Row_Counter+=1
    if Save_counter == 50:
        comp.save('Scrap.xlsx')
        print('Proccesing Now ...\n',i,'From',sheet.max_row)
        Save_counter=0


            